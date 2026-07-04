"""
Bank Statement Categorization — Single-Port FastAPI App
Serves the React frontend (from ../frontend/dist) on the same port as the API.
"""
import io
import os
from pathlib import Path
from typing import List, Optional

import pandas as pd
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from sqlalchemy.orm import Session

import crud
import models
import schemas
from database import engine, get_db, run_migrations
from pdf_parser import parse_ktb_pdf
from ttb_parser import parse_ttb_pdf

# ─── App Setup ───────────────────────────────────────────────────────────────

models.Base.metadata.create_all(bind=engine)

app = FastAPI(
    title="Bank Statement Categorization API",
    version="2.0.0",
    # Serve docs at /api/docs so they don't conflict with the SPA
    docs_url="/api/docs",
    redoc_url="/api/redoc",
    openapi_url="/api/openapi.json",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Startup ─────────────────────────────────────────────────────────────────

@app.on_event("startup")
def startup_event():
    run_migrations()
    db = next(get_db())
    crud.seed_default_categories(db)


# ─── API Routes (all prefixed /api) ──────────────────────────────────────────

# Health
@app.get("/api/health", tags=["Health"])
def health():
    return {"status": "ok"}


# ─── Sessions ────────────────────────────────────────────────────────────────

@app.get("/api/sessions", response_model=List[schemas.UploadSessionOut], tags=["Sessions"])
def list_sessions(db: Session = Depends(get_db)):
    return crud.get_sessions(db)


@app.get("/api/sessions/{session_id}", response_model=schemas.UploadSessionOut, tags=["Sessions"])
def get_session(session_id: int, db: Session = Depends(get_db)):
    s = crud.get_session(db, session_id)
    if not s:
        raise HTTPException(404, "Session not found")
    return s


@app.delete("/api/sessions/{session_id}", tags=["Sessions"])
def delete_session(session_id: int, db: Session = Depends(get_db)):
    crud.delete_session(db, session_id)
    return {"message": "Session and transactions deleted"}


# ─── Transactions ─────────────────────────────────────────────────────────────

@app.get("/api/transactions", response_model=List[schemas.TransactionOut], tags=["Transactions"])
def get_transactions(
    session_id: Optional[int] = Query(None),
    session_ids: Optional[str] = Query(None),  # comma-separated: "1,2,3"
    skip: int = Query(0, ge=0),
    limit: int = Query(50000, ge=1),
    db: Session = Depends(get_db),
):
    ids = None
    if session_ids:
        ids = [int(x) for x in session_ids.split(",") if x.strip().isdigit()]
    return crud.get_transactions(db, session_id=session_id, session_ids=ids, skip=skip, limit=limit)


@app.get("/api/transactions/stats", tags=["Transactions"])
def get_stats(session_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    return crud.get_transaction_stats(db, session_id=session_id)


@app.get("/api/transactions/category-summary", tags=["Transactions"])
def get_category_summary(session_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    return crud.get_category_summary(db, session_id=session_id)


@app.put("/api/transactions/bulk", tags=["Transactions"])
def bulk_update(request: schemas.BulkUpdateRequest, db: Session = Depends(get_db)):
    updated = crud.bulk_update_transactions(
        db, ids=request.ids, category=request.category, status=request.status
    )
    session_ids = set(
        tx.session_id for tx in
        db.query(models.Transaction).filter(models.Transaction.id.in_(request.ids)).all()
        if tx.session_id
    )
    for sid in session_ids:
        crud.update_session_counts(db, sid)
    return {"updated_count": len(updated), "message": f"Updated {len(updated)} transactions"}


@app.get("/api/transactions/distinct-categories", tags=["Transactions"])
def get_distinct_categories(db: Session = Depends(get_db)):
    """คืน list ของชื่อ category ที่มีอยู่ใน transactions จริง พร้อมจำนวน"""
    from sqlalchemy import func as sqlfunc
    rows = (
        db.query(models.Transaction.category, sqlfunc.count(models.Transaction.id).label("count"))
        .group_by(models.Transaction.category)
        .order_by(models.Transaction.category)
        .all()
    )
    return [{"category": r.category, "count": r.count} for r in rows]


@app.post("/api/transactions/remap-category", tags=["Transactions"])
def remap_category(body: dict, db: Session = Depends(get_db)):
    """เปลี่ยนชื่อ category เก่า → ใหม่ ใน transactions ทั้งหมด"""
    old_name = body.get("old_name", "").strip()
    new_name = body.get("new_name", "").strip()
    if not old_name or not new_name:
        raise HTTPException(400, "old_name และ new_name ต้องไม่ว่าง")
    updated = (
        db.query(models.Transaction)
        .filter(models.Transaction.category == old_name)
        .update({"category": new_name}, synchronize_session=False)
    )
    db.commit()
    # อัปเดต session counts ทุก session ที่ได้รับผลกระทบ
    affected_sessions = (
        db.query(models.Transaction.session_id)
        .filter(models.Transaction.category == new_name)
        .distinct()
        .all()
    )
    for (sid,) in affected_sessions:
        if sid:
            crud.update_session_counts(db, sid)
    return {"updated_count": updated, "old_name": old_name, "new_name": new_name}


@app.put("/api/transactions/{transaction_id}", response_model=schemas.TransactionOut, tags=["Transactions"])
def update_transaction(
    transaction_id: int,
    update: schemas.TransactionUpdate,
    db: Session = Depends(get_db),
):
    tx = crud.update_transaction(db, transaction_id, update)
    if not tx:
        raise HTTPException(404, "Transaction not found")
    if tx.session_id:
        crud.update_session_counts(db, tx.session_id)
    return tx


@app.delete("/api/transactions", tags=["Transactions"])
def delete_all(db: Session = Depends(get_db)):
    crud.delete_all_transactions(db)
    return {"message": "All data deleted"}


# ─── Upload PDF ───────────────────────────────────────────────────────────────

@app.post("/api/upload", tags=["Upload"])
async def upload_file(
    file: UploadFile = File(...),
    bank: str = Query("ktb", description="ktb | ttb"),
    db: Session = Depends(get_db)
):
    filename = file.filename or ""
    ext = Path(filename).suffix.lower()

    if ext not in (".pdf", ".xlsx", ".xls", ".csv", ".json"):
        raise HTTPException(400, "Supported formats: PDF, Excel (.xlsx/.xls), CSV, JSON")

    contents = await file.read()

    if ext == ".pdf":
        return await _handle_pdf_upload(contents, filename, db, bank=bank)
    elif ext == ".json":
        return await _handle_json_upload(contents, filename, db)
    else:
        return await _handle_excel_upload(contents, filename, ext, db)


async def _handle_pdf_upload(contents: bytes, filename: str, db: Session, bank: str = "ktb"):
    import hashlib
    file_hash = hashlib.md5(contents).hexdigest()

    existing = crud.get_session_by_hash(db, file_hash)
    if existing:
        return {
            "message": f"ไฟล์นี้เคยนำเข้าแล้ว (session #{existing.id}) — กำลังโหลดข้อมูลเดิม",
            "session_id": existing.id,
            "count": existing.total_rows,
            "is_duplicate": True,
        }

    try:
        if bank == "ttb":
            statement = parse_ttb_pdf(contents, filename)
        else:
            statement = parse_ktb_pdf(contents, filename)
    except Exception as e:
        raise HTTPException(400, f"ไม่สามารถอ่าน PDF ได้: {str(e)}")

    if not statement.transactions:
        raise HTTPException(400, "ไม่พบข้อมูลรายการในไฟล์ PDF — กรุณาตรวจสอบรูปแบบไฟล์")

    # Create session
    session = crud.create_session(
        db,
        filename=filename,
        file_hash=file_hash,
        account_number=statement.account_number,
        account_name=statement.account_name,
        bank_name=statement.bank_name,
        period_start=statement.period_start,
        period_end=statement.period_end,
    )

    # Import transactions
    tx_creates = [
        schemas.TransactionCreate(
            session_id=session.id,
            date=tx.date,
            particulars=tx.particulars,
            chq_no=tx.chq_no,
            withdrawal=tx.withdrawal,
            deposit=tx.deposit,
            balance=tx.balance,
            via=tx.via,
        )
        for tx in statement.transactions
    ]
    crud.bulk_create_transactions(db, tx_creates)
    crud.update_session_counts(db, session.id)

    return {
        "message": f"นำเข้า {len(tx_creates)} รายการจาก {filename} สำเร็จ",
        "session_id": session.id,
        "count": len(tx_creates),
        "account_number": statement.account_number,
        "account_name": statement.account_name,
        "period_start": statement.period_start,
        "period_end": statement.period_end,
        "is_duplicate": False,
    }


async def _handle_json_upload(contents: bytes, filename: str, db: Session):
    """Parse JSON array and import transactions.

    รองรับ format:
    [
      {
        "date": "13/05/24",          ← DD/MM/YY หรือ DD/MM/YYYY
        "tx_amount": "373.00",       ← จำนวนเงิน (บวก = ฝาก, ลบ = ถอน)
        "balance": "6,751.57",
        "detail": "TR 9906901",      ← รายละเอียดรายการ
        "category": "ลูกค้าโอนให้"   ← optional: เก็บไว้ตามที่มีแม้ไม่อยู่ใน DB
      },
      ...
    ]
    """
    import json
    import hashlib

    file_hash = hashlib.md5(contents).hexdigest()

    # ตรวจซ้ำ
    existing = crud.get_session_by_hash(db, file_hash)
    if existing:
        return {
            "message": f"ไฟล์นี้เคยนำเข้าแล้ว (session #{existing.id}) — กำลังโหลดข้อมูลเดิม",
            "session_id": existing.id,
            "count": existing.total_rows,
            "is_duplicate": True,
        }

    try:
        data = json.loads(contents.decode("utf-8-sig"))
    except Exception as e:
        raise HTTPException(400, f"อ่าน JSON ไม่ได้: {str(e)}")

    if not isinstance(data, list):
        raise HTTPException(400, "ไฟล์ JSON ต้องเป็น array ของ object")

    if len(data) == 0:
        raise HTTPException(400, "ไม่พบข้อมูลใน JSON")

    def clean_number(v) -> Optional[float]:
        """แปลง '6,751.57' หรือ '-373.00' → float"""
        if v is None:
            return None
        try:
            return float(str(v).replace(",", "").strip())
        except (ValueError, TypeError):
            return None

    def parse_date_sort_key(date_str: str) -> str:
        """แปลง DD/MM/YY หรือ DD/MM/YYYY → YYYY-MM-DD สำหรับเรียงลำดับ"""
        if not date_str:
            return ""
        parts = date_str.strip().split("/")
        if len(parts) == 3:
            day, month, year = parts
            # DD/MM/YY → เติม 20xx
            if len(year) == 2:
                year = "20" + year
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        return date_str

    session = crud.create_session(db, filename=filename, file_hash=file_hash)

    # เรียงข้อมูลตามวันที่ก่อน import
    try:
        data = sorted(data, key=lambda r: parse_date_sort_key(str(r.get("date", ""))))
    except Exception:
        pass  # ถ้าเรียงไม่ได้ ใช้ลำดับเดิม

    transactions = []
    for row in data:
        if not isinstance(row, dict):
            continue

        # รองรับทั้ง "detail" และ "particulars"
        particulars = row.get("detail") or row.get("particulars") or row.get("description")

        # tx_amount: ถ้าเป็นบวก = ฝาก, ถ้าเป็นลบ = ถอน
        raw_amount = clean_number(row.get("tx_amount") or row.get("amount"))
        withdrawal = None
        deposit = None
        if raw_amount is not None:
            if raw_amount < 0:
                withdrawal = abs(raw_amount)
            else:
                deposit = raw_amount

        # รองรับ withdrawal/deposit แยกกัน (ถ้ามี)
        if row.get("withdrawal") is not None:
            withdrawal = clean_number(row.get("withdrawal"))
        if row.get("deposit") is not None:
            deposit = clean_number(row.get("deposit"))

        # เก็บ category ตามที่มีใน JSON เลย ไม่ fallback เป็น Uncategorized
        # เพื่อให้ผู้ใช้เห็นข้อมูลจริงและ remap ได้ภายหลัง
        category = str(row.get("category", "")).strip() or "Uncategorized"

        tx = schemas.TransactionCreate(
            session_id=session.id,
            date=str(row.get("date", "")).strip() or None,
            particulars=str(particulars).strip() if particulars else None,
            chq_no=str(row.get("chq_no", "")).strip() or None,
            withdrawal=withdrawal,
            deposit=deposit,
            balance=clean_number(row.get("balance")),
            via=str(row.get("via", "")).strip() or None,
            category=category,
        )
        transactions.append(tx)

    if not transactions:
        crud.delete_session(db, session.id)
        raise HTTPException(400, "ไม่พบข้อมูลรายการใน JSON ที่สามารถนำเข้าได้")

    crud.bulk_create_transactions(db, transactions)
    crud.update_session_counts(db, session.id)

    return {
        "message": f"นำเข้า {len(transactions)} รายการจาก {filename} สำเร็จ",
        "session_id": session.id,
        "count": len(transactions),
        "is_duplicate": False,
    }


COLUMN_ALIASES = {    "date": "date", "วันที่": "date",
    "particulars": "particulars", "รายการ": "particulars", "description": "particulars",
    "withdrawal": "withdrawal", "ถอน": "withdrawal", "debit": "withdrawal",
    "deposit": "deposit", "ฝาก": "deposit", "credit": "deposit",
    "balance": "balance", "คงเหลือ": "balance", "ยอดคงเหลือ": "balance",
    "via": "via", "ช่องทาง": "via", "channel": "via",
}


async def _handle_excel_upload(contents: bytes, filename: str, ext: str, db: Session):
    """Parse Excel/CSV and import transactions."""
    try:
        if ext == ".csv":
            df = pd.read_csv(io.BytesIO(contents), encoding="utf-8-sig")
        else:
            df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(400, f"อ่านไฟล์ไม่ได้: {str(e)}")

    df.columns = [str(c).strip() for c in df.columns]
    df.rename(columns={col: COLUMN_ALIASES[col] for col in df.columns if col in COLUMN_ALIASES}, inplace=True)

    key_cols = [c for c in ["date", "particulars", "withdrawal", "deposit", "balance"] if c in df.columns]
    df.dropna(how="all", subset=key_cols or None, inplace=True)

    if df.empty:
        raise HTTPException(400, "ไม่พบข้อมูลในไฟล์")

    session = crud.create_session(db, filename=filename)

    def safe_str(v): return str(v).strip() if pd.notna(v) else None
    def safe_float(v):
        try: return float(v) if pd.notna(v) else None
        except: return None

    transactions = [
        schemas.TransactionCreate(
            session_id=session.id,
            date=safe_str(row.get("date")),
            particulars=safe_str(row.get("particulars")),
            withdrawal=safe_float(row.get("withdrawal")),
            deposit=safe_float(row.get("deposit")),
            balance=safe_float(row.get("balance")),
            via=safe_str(row.get("via")),
        )
        for _, row in df.iterrows()
    ]

    crud.bulk_create_transactions(db, transactions)
    crud.update_session_counts(db, session.id)

    return {
        "message": f"นำเข้า {len(transactions)} รายการสำเร็จ",
        "session_id": session.id,
        "count": len(transactions),
        "is_duplicate": False,
    }


# ─── Export ───────────────────────────────────────────────────────────────────

CATEGORY_FILL_COLORS = {
    "จ่ายพนักงาน":    "FFCDD2",
    "ซื้อของ":         "FFE0B2",
    "ค่าสาธารณูปโภค": "FFF9C4",
    "รายได้":          "C8E6C9",
    "โอนเงิน":         "BBDEFB",
    "ค่าใช้จ่ายทั่วไป":"E1BEE7",
    "Uncategorized":   "F5F5F5",
}

# Ordered export sheets — maps sheet name → list of categories to include
EXPORT_SHEET_MAP = [
    ("โอนออกซื้อของ",          ["ซื้อของ"]),
    ("โอนออกแยกบัญชีตัวเอง",   ["โอนเงิน"]),
    ("จ่ายพนักงาน",             ["จ่ายพนักงาน"]),
    ("ลูกค้าโอนให้",            ["รายได้"]),
    ("ค่าสาธารณูปโภค",          ["ค่าสาธารณูปโภค"]),
    ("อื่นๆ",                   ["ค่าใช้จ่ายทั่วไป", "Uncategorized"]),
]


def _style_header(ws, num_cols: int):
    """Apply dark header style to row 1."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    hdr_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 28


def _style_data_rows(ws, transactions_slice, col_count: int, fill_hex: str = None):
    """Apply fill + border to data rows. If fill_hex is None, uses per-row category color."""
    from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for row_idx, tx in enumerate(transactions_slice, start=2):
        hex_color = fill_hex or CATEGORY_FILL_COLORS.get(tx.category, "FFFFFF")
        row_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.border = thin
            cell.alignment = Alignment(vertical="center")


def _set_col_widths(ws, df):
    from openpyxl.utils import get_column_letter
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_data = df.iloc[:, col_idx - 1]
        try:
            str_lens = col_data.apply(lambda x: len(str(x)) if x is not None and str(x) != 'nan' else 0)
            max_data_len = int(str_lens.max()) if len(str_lens) > 0 else 0
        except Exception:
            max_data_len = 0
        max_len = max(len(str(col_name)), max_data_len)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)


def _style_total_row(ws, row_idx: int, col_count: int):
    """Style subtotal row (dark background, bold white)."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    total_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF", size=10)
    med_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="medium"), bottom=Side(style="medium"),
    )
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = med_border
        cell.alignment = Alignment(
            horizontal="right" if col_idx >= 5 else "left",
            vertical="center",
        )
        if col_idx in (5, 6):
            cell.number_format = '#,##0.00'


@app.get("/api/export", tags=["Export"])
def export_excel(
    session_id: Optional[int] = Query(None),
    session_ids: Optional[str] = Query(None),  # comma-separated: "1,2,3"
    db: Session = Depends(get_db)
):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ids = None
    if session_ids:
        ids = [int(x) for x in session_ids.split(",") if x.strip().isdigit()]

    transactions = crud.get_transactions(db, session_id=session_id, session_ids=ids)
    if not transactions:
        raise HTTPException(404, "ไม่มีข้อมูลสำหรับ Export")

    db_cats = crud.get_categories(db)
    db_cat_color = {c.name: c.color.lstrip("#") for c in db_cats}

    # ถ้าเลือกหลาย session ให้ชื่อไฟล์เป็น combined
    if ids and len(ids) > 1:
        filename_base = f"combined_{len(ids)}_files"
    elif session_id:
        session_info = crud.get_session(db, session_id)
        filename_base = Path(session_info.filename).stem if session_info else "bank_statement"
    elif ids and len(ids) == 1:
        session_info = crud.get_session(db, ids[0])
        filename_base = Path(session_info.filename).stem if session_info else "bank_statement"
    else:
        filename_base = "bank_statement"

    output = io.BytesIO()

    # คอลัมน์ใน sheet รายการ (ลำดับสำคัญมากสำหรับ formula อ้างอิง)
    # E=ถอน(฿), F=ฝาก(฿), H=หมวดหมู่
    COLS = ["วันที่", "รายการ", "ผ่านทาง", "เลขที่เช็ค", "ถอน (฿)", "ฝาก (฿)", "คงเหลือ (฿)", "หมวดหมู่"]
    COL_WD  = 5   # column index (1-based) ของ ถอน (฿)
    COL_DEP = 6   # column index ของ ฝาก (฿)
    COL_CAT = 8   # column index ของ หมวดหมู่

    def tx_to_row(tx):
        return {
            "วันที่":      tx.date,
            "รายการ":      tx.particulars,
            "ผ่านทาง":     tx.via,
            "เลขที่เช็ค":  tx.chq_no,
            "ถอน (฿)":     tx.withdrawal,
            "ฝาก (฿)":     tx.deposit,
            "คงเหลือ (฿)": tx.balance,
            "หมวดหมู่":    tx.category,
        }

    # helper: ชื่อ column letter
    WD_COL  = get_column_letter(COL_WD)
    DEP_COL = get_column_letter(COL_DEP)
    CAT_COL = get_column_letter(COL_CAT)

    def _hex_to_fill(hex6: str) -> PatternFill:
        """สร้าง PatternFill จาก hex string 6 ตัว"""
        return PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")

    def _quote_sheet(name: str) -> str:
        """ครอบชื่อ sheet ด้วย single quote สำหรับ formula reference"""
        return f"'{name}'"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        wb = writer.book

        # ── Sheet: รายการทั้งหมด ────────────────────────────────────────────
        all_rows = [tx_to_row(tx) for tx in transactions]
        df_all = pd.DataFrame(all_rows, columns=COLS)
        df_all.to_excel(writer, index=False, sheet_name="รายการทั้งหมด")
        ws_all = writer.sheets["รายการทั้งหมด"]
        _style_header(ws_all, len(COLS))
        _style_data_rows(ws_all, transactions, len(COLS))
        _set_col_widths(ws_all, df_all)
        ws_all.freeze_panes = "A2"
        # สูตร subtotal row ท้าย sheet รายการทั้งหมด
        sub_row = len(transactions) + 2
        n_data  = len(transactions)
        ws_all.cell(sub_row, 1, "รวม").font = Font(bold=True)
        ws_all.cell(sub_row, 2, f"{n_data} รายการ")
        ws_all.cell(sub_row, COL_WD,  f"=SUM({WD_COL}2:{WD_COL}{sub_row-1})")
        ws_all.cell(sub_row, COL_DEP, f"=SUM({DEP_COL}2:{DEP_COL}{sub_row-1})")
        _style_total_row(ws_all, sub_row, len(COLS))

        # ── Sheets per category group ────────────────────────────────────────
        cat_map: dict[str, list] = {}
        for tx in transactions:
            cat_map.setdefault(tx.category, []).append(tx)

        covered_cats = set()
        for _, cats in EXPORT_SHEET_MAP:
            covered_cats.update(cats)
        extra_cats = [c for c in cat_map if c not in covered_cats]

        sheet_list = list(EXPORT_SHEET_MAP)
        for cat in extra_cats:
            sheet_list.append((cat[:31], [cat]))

        # เก็บ mapping sheet_name → (list_of_category_names, data_row_count)
        # สำหรับใช้สร้าง formula ใน summary sheet
        sheet_meta: dict[str, dict] = {}

        for sheet_name, cat_names in sheet_list:
            group_txs = []
            for cat in cat_names:
                group_txs.extend(cat_map.get(cat, []))
            if not group_txs:
                continue

            rows = [tx_to_row(tx) for tx in group_txs]
            df_g = pd.DataFrame(rows, columns=COLS)
            df_g.to_excel(writer, index=False, sheet_name=sheet_name)

            ws = writer.sheets[sheet_name]
            _style_header(ws, len(COLS))
            _style_data_rows(ws, group_txs, len(COLS))
            _set_col_widths(ws, df_g)
            ws.freeze_panes = "A2"

            # subtotal row ด้วย SUM formula
            sub_r = len(group_txs) + 2
            ws.cell(sub_r, 1, "รวม").font = Font(bold=True)
            ws.cell(sub_r, 2, f"{len(group_txs)} รายการ")
            ws.cell(sub_r, COL_WD,  f"=SUM({WD_COL}2:{WD_COL}{sub_r-1})")
            ws.cell(sub_r, COL_DEP, f"=SUM({DEP_COL}2:{DEP_COL}{sub_r-1})")
            _style_total_row(ws, sub_r, len(COLS))

            sheet_meta[sheet_name] = {
                "cat_names": cat_names,
                "data_rows": len(group_txs),
                "sub_row":   sub_r,
            }

        # ── Summary sheet (formula-based) ────────────────────────────────────
        ws_s = wb.create_sheet("สรุปยอดรวม")

        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # Account info block
        row = 1
        if session_info:
            info_rows = [
                ("ไฟล์ต้นฉบับ", session_info.filename),
                ("เลขบัญชี",    session_info.account_number or "-"),
                ("ชื่อบัญชี",   session_info.account_name   or "-"),
                ("ธนาคาร",      session_info.bank_name      or "-"),
                ("ช่วงเวลา",    f"{session_info.period_start} – {session_info.period_end}"),
            ]
            for label, val in info_rows:
                ws_s.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
                ws_s.cell(row=row, column=2, value=val)
                row += 1
            row += 1  # blank line

        # Header row
        sum_headers = ["หมวดหมู่", "จำนวนรายการ", "ยอดถอนรวม (฿)", "ยอดฝากรวม (฿)", "สุทธิ (฿)"]
        hdr_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        for col_idx, h in enumerate(sum_headers, 1):
            cell = ws_s.cell(row=row, column=col_idx, value=h)
            cell.fill = hdr_fill
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin
        ws_s.row_dimensions[row].height = 24
        header_row = row
        row += 1

        # Build ordered category list (same order as before)
        ordered = []
        seen = set()
        for _, cats in EXPORT_SHEET_MAP:
            for c in cats:
                if c in cat_map and c not in seen:
                    ordered.append(c)
                    seen.add(c)
        for c in cat_map:
            if c not in seen:
                ordered.append(c)

        # data_row_start ใน summary (สำหรับ SUM formula grand total)
        data_start_row = row
        cat_row_map: dict[str, int] = {}   # category → row ใน summary

        for cat in ordered:
            # หา sheet ที่ cat อยู่
            target_sheet = None
            for sname, meta in sheet_meta.items():
                if cat in meta["cat_names"]:
                    target_sheet = sname
                    break

            ref = _quote_sheet(target_sheet) if target_sheet else None
            n_txs = len(cat_map.get(cat, []))

            # สร้าง formula อ้างอิงไป sheet นั้น
            # SUMIF เพื่อกันบรรทัด subtotal "รวม" ไม่ให้นับ
            if ref and target_sheet:
                sub_r = sheet_meta[target_sheet]["sub_row"]
                data_end = sub_r - 1
                # criteria ใน SUMIF/COUNTIF ต้องใช้ double quote ครอบชื่อหมวดหมู่
                cat_escaped = cat.replace('"', '""')   # escape double quote ในชื่อ
                wd_formula  = f'=SUMIF({ref}!{CAT_COL}2:{CAT_COL}{data_end},"{cat_escaped}",{ref}!{WD_COL}2:{WD_COL}{data_end})'
                dep_formula = f'=SUMIF({ref}!{CAT_COL}2:{CAT_COL}{data_end},"{cat_escaped}",{ref}!{DEP_COL}2:{DEP_COL}{data_end})'
                cnt_formula = f'=COUNTIF({ref}!{CAT_COL}2:{CAT_COL}{data_end},"{cat_escaped}")'
            else:
                # fallback hardcode ถ้าไม่มี sheet
                txs = cat_map.get(cat, [])
                wd_formula  = sum(t.withdrawal or 0 for t in txs)
                dep_formula = sum(t.deposit    or 0 for t in txs)
                cnt_formula = n_txs

            cat_col_letter = "C"   # ยอดถอน
            dep_col_letter = "D"   # ยอดฝาก

            # fill สี
            hex6 = db_cat_color.get(cat, CATEGORY_FILL_COLORS.get(cat, "FFFFFF"))
            row_fill = _hex_to_fill(hex6)

            # เขียน row
            for col_idx in range(1, 6):
                cell = ws_s.cell(row=row, column=col_idx)
                cell.fill = row_fill
                cell.border = thin
                cell.alignment = Alignment(
                    horizontal="right" if col_idx >= 2 else "left",
                    vertical="center",
                )

            ws_s.cell(row=row, column=1, value=cat)
            ws_s.cell(row=row, column=2, value=cnt_formula)
            ws_s.cell(row=row, column=3, value=wd_formula)
            ws_s.cell(row=row, column=4, value=dep_formula)
            # สุทธิ = ฝาก - ถอน (formula อ้างอิงเซลล์เดียวกัน)
            ws_s.cell(row=row, column=5, value=f"=D{row}-C{row}")

            # format number
            for col_idx in [3, 4, 5]:
                ws_s.cell(row=row, column=col_idx).number_format = '#,##0.00'

            cat_row_map[cat] = row
            row += 1

        data_end_row = row - 1

        # Grand total row — SUM formula รวม rows ด้านบน
        grand_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
        grand_font = Font(bold=True, color="FFFFFF", size=10)
        grand_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="medium"), bottom=Side(style="medium"),
        )
        grand_data = [
            ("รวมทั้งหมด",          None),
            (f"=SUM(B{data_start_row}:B{data_end_row})", None),
            (f"=SUM(C{data_start_row}:C{data_end_row})", '#,##0.00'),
            (f"=SUM(D{data_start_row}:D{data_end_row})", '#,##0.00'),
            (f"=D{row}-C{row}",                          '#,##0.00'),
        ]
        for col_idx, (val, fmt) in enumerate(grand_data, 1):
            cell = ws_s.cell(row=row, column=col_idx, value=val)
            cell.fill = grand_fill
            cell.font = grand_font
            cell.border = grand_border
            cell.alignment = Alignment(
                horizontal="right" if col_idx >= 2 else "left",
                vertical="center",
            )
            if fmt:
                cell.number_format = fmt

        # Column widths
        ws_s.column_dimensions["A"].width = 28
        ws_s.column_dimensions["B"].width = 15
        ws_s.column_dimensions["C"].width = 20
        ws_s.column_dimensions["D"].width = 20
        ws_s.column_dimensions["E"].width = 20

        # Move summary to front
        wb.move_sheet("สรุปยอดรวม", offset=-len(wb.sheetnames) + 1)

    output.seek(0)
    safe_name = f"{filename_base}_categorized.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )


# ─── Categories ───────────────────────────────────────────────────────────────

@app.get("/api/categories", response_model=List[schemas.CategoryConfigOut], tags=["Categories"])
def get_categories(db: Session = Depends(get_db)):
    return crud.get_categories(db)


@app.post("/api/categories", response_model=schemas.CategoryConfigOut, tags=["Categories"])
def create_category(category: schemas.CategoryConfigCreate, db: Session = Depends(get_db)):
    return crud.create_category(db, category)


@app.put("/api/categories/reorder", response_model=List[schemas.CategoryConfigOut], tags=["Categories"])
def reorder_categories(request: schemas.CategoryReorderRequest, db: Session = Depends(get_db)):
    return crud.reorder_categories(db, request.items)


@app.put("/api/categories/{category_id}", response_model=schemas.CategoryConfigOut, tags=["Categories"])
def update_category(category_id: int, update: schemas.CategoryConfigUpdate, db: Session = Depends(get_db)):
    cat = crud.update_category(db, category_id, update)
    if not cat:
        raise HTTPException(404, "Category not found")
    return cat


@app.delete("/api/categories/{category_id}", tags=["Categories"])
def delete_category(category_id: int, db: Session = Depends(get_db)):
    cat = crud.delete_category(db, category_id)
    if not cat:
        raise HTTPException(404, "Category not found")
    return {"message": "Category deleted"}


# ─── Serve React SPA (must be LAST) ──────────────────────────────────────────

FRONTEND_DIST = Path(__file__).parent.parent / "frontend" / "dist"

if FRONTEND_DIST.exists():
    app.mount("/assets", StaticFiles(directory=str(FRONTEND_DIST / "assets")), name="assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    def serve_spa(full_path: str):
        # Don't serve SPA for /api routes (already handled above)
        if full_path.startswith("api/"):
            raise HTTPException(404)
        index = FRONTEND_DIST / "index.html"
        if index.exists():
            return FileResponse(str(index))
        raise HTTPException(404, "Frontend not built. Run: cd frontend && npm run build")

    @app.get("/", include_in_schema=False)
    def serve_root():
        index = FRONTEND_DIST / "index.html"
        if index.exists():
            return FileResponse(str(index))
        return {"message": "API is running. Frontend not built yet.", "docs": "/api/docs"}
